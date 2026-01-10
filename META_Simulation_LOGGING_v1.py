import pandas as pd
from pathlib import Path

# ======================================================================================
# KPI FUNCTIONS (UNCHANGED) 
# ======================================================================================

# is a function to return the detailed structure of the current situation of the storage: {article_code, crate_type, quantity}
def kpi_storage_detail(storage):
    """
    Returns a detailed overview of all crates in the storage area.

    Output format:
    {
        article_code: {
            crate_type: quantity
        }
    }
    """
    result = {}

    for (article_code, crate_type), quantity in storage.items():
        if article_code not in result:
            result[article_code] = {}

        result[article_code][crate_type] = quantity

    return result

# is a function to return the total n° of crates in the current storage
def kpi_total_crates(storage):
    return sum(storage.values())

# ======================================================================================
# STORAGE LOGGING
# ======================================================================================

from pathlib import Path
from datetime import datetime
from openpyxl import Workbook, load_workbook
from META_Simulation_CONFIG_v1 import get_crate_family


def kpi_family_counts(storage):
    """
    Returns totals by family based on current storage.
    Output: (ifco_total, customer_total)
    """
    ifco_crates = 0
    customer_crates = 0

    for (article_code, crate_type), qty in storage.items():
        family = get_crate_family(crate_type)
        if family == "IFCO":
            ifco_crates += qty
        elif family == "CUSTOMER_TOTE":
            customer_crates += qty

    return ifco_crates, customer_crates


class IterationExcelLogger:
    def __init__(self, filepath="Storage_Log.xlsx", sheet_name="log"):
        self.filepath = Path(filepath)
        self.sheet_name = sheet_name
        self._ensure_file()

    def _ensure_file(self):
        # Create Excel file with headers if it doesn't exist
        if self.filepath.exists():
            return

        wb = Workbook()
        ws = wb.active
        ws.title = self.sheet_name

        ws.append([
            "Iteration",
            "infeed_until_now",           # cumulative sum of crates fed in
            "total_crates_in_storage_now", # sum of ifco + customer in storage
            "ifco_crates_in_storage",     # snapshot
            "customer_crates_in_storage", # snapshot
        ])

        wb.save(self.filepath)

    def append_row(self, iteration, infeed_until_now, storage):
        ifco_crates, customer_crates = kpi_family_counts(storage)
        total_now = ifco_crates + customer_crates

        wb = load_workbook(self.filepath)
        ws = wb[self.sheet_name]

        ws.append([
            int(iteration),
            int(infeed_until_now),
            int(total_now),
            int(ifco_crates),
            int(customer_crates)
        ])


        wb.save(self.filepath)


# ======================================================================================
# INFEED LOGGING
# ======================================================================================

def log_infeed_iteration(
    iteration,
    infeed_obj=None,
    file_path="Infeed_Log.xlsx"
):
    """
    Logs exactly one row per iteration into an Excel file.

    Columns:
    - Iteration
    - article_code
    - crate_type
    - IN-quantity

    If no infeed happened, placeholders are written.
    """

    # Prepare row data
    if infeed_obj is not None:
        row = {
            "Iteration": iteration,
            "article_code": infeed_obj.article_code,
            "crate_type": infeed_obj.crate_type,
            "IN-quantity": infeed_obj.quantity
        }
    else:
        row = {
            "Iteration": iteration,
            "article_code": "",
            "crate_type": "",
            "IN-quantity": ""
        }

    df_new = pd.DataFrame([row])

    file = Path(file_path)

    if file.exists():
        df_existing = pd.read_excel(file)
        df_combined = pd.concat([df_existing, df_new], ignore_index=True)
    else:
        df_combined = df_new

    df_combined.to_excel(file, index=False)

# ======================================================================================
# OUTFEED LOGGING
# ======================================================================================

import pandas as pd
from pathlib import Path

from META_Simulation_CONFIG_v1 import (
    get_stack_family,
    STACKS_PER_PALLET
)
from META_Simulation_OUTFEED_v3 import calculate_stack_height


def log_outfeed_iteration(
    iteration,
    pallet_built,
    shop=None,
    stacks=None,
    file_path="Outfeed_Log.xlsx"
):
    """
    Logs exactly one row per iteration for outfeed activity.

    One row is written per iteration, even if no pallet was built.
    """

    # Base columns
    row = {
        "iteration": iteration,
        "pallet_built": int(pallet_built),
        "shop": shop if pallet_built else "",
        "total_crates_pallet": "",
    }

    # Initialize stack-specific columns
    for i in range(1, 5):
        row[f"stack_height_{i}"] = ""
        row[f"stack_family_{i}"] = ""
        row[f"crate_qty_{i}"] = ""
    
    # Fill data if pallet was built
    if pallet_built and len(stacks) == STACKS_PER_PALLET:
        total_crates = 0

        for idx, stack in enumerate(stacks):
            if idx >= 4:
                break

            if stack:
                stack_height = calculate_stack_height(stack)
                stack_family = get_stack_family(stack)
                crate_qty = sum(qty for _, _, qty in stack)

                row[f"stack_height_{idx + 1}"] = int(stack_height)
                row[f"stack_family_{idx + 1}"] = stack_family
                row[f"crate_qty_{idx + 1}"] = int(crate_qty)

                total_crates += crate_qty

        row["total_crates_pallet"] = int(total_crates)

    # Write to Excel
    df_new = pd.DataFrame([row])
    file = Path(file_path)

    if file.exists():
        df_existing = pd.read_excel(file)
        df_combined = pd.concat([df_existing, df_new], ignore_index=True)
    else:
        df_combined = df_new

    df_combined.to_excel(file, index=False)
